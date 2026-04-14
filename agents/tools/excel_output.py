# agents/tools/excel_output.py
from pathlib import Path
from typing import Any
from pydantic import BaseModel, Field
from crewai.tools import BaseTool
from api.config import get_settings
from agents.tools._db import insert_agent_output_sync


class ExcelOutputToolInput(BaseModel):
    rows: list[dict[str, Any]] = Field(
        description="List of dicts representing rows. All dicts must share the same keys."
    )
    columns: list[str] = Field(
        description="Ordered list of column names to include in the output."
    )
    filename: str = Field(
        description="Output filename (e.g. 'portfolio_register.xlsx'). "
        ".xlsx extension added automatically if missing."
    )
    agent_name: str = Field(
        description="Name of the agent producing this output (used for output tracking)."
    )


class ExcelOutputTool(BaseTool):
    name: str = "ExcelOutputTool"
    description: str = (
        "Write a list of records to an Excel (.xlsx) file in the project outputs directory. "
        "Pass rows as a list of dicts with uniform keys, an ordered column list, and a filename. "
        "Returns the absolute file path to the saved file."
    )
    args_schema: type[BaseModel] = ExcelOutputToolInput
    slug: str

    def _run(
        self,
        rows: list[dict[str, Any]],
        columns: list[str],
        filename: str,
        agent_name: str,
    ) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            return "Error: openpyxl not installed — run: pip install openpyxl"

        settings = get_settings()
        outputs_dir = Path(settings.projects_dir) / self.slug / "outputs"
        outputs_dir.mkdir(parents=True, exist_ok=True)

        if not filename.endswith(".xlsx"):
            filename = f"{filename}.xlsx"
        file_path = outputs_dir / filename

        try:
            wb = openpyxl.Workbook()
            ws = wb.active

            # Header row — bold
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)

            # Data rows
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

            # Auto-width columns (capped at 60 to avoid absurdly wide columns)
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            # Freeze header row
            ws.freeze_panes = "A2"

            wb.save(file_path)
            insert_agent_output_sync(
                slug=self.slug,
                agent_name=agent_name,
                output_type="excel",
                file_path=str(file_path),
            )
        except (OSError, ValueError) as e:
            return f"Error: write failed — {e}"

        return str(file_path)
