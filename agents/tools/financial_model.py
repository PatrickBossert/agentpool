# agents/tools/financial_model.py
from pathlib import Path
from pydantic import BaseModel, Field
from crewai.tools import BaseTool
from api.config import get_settings
from agents.tools._db import insert_agent_output_sync


def _calculate_npv(cashflows: list[float], rate_per_period: float) -> float:
    """Discounted NPV across all cashflow periods."""
    return sum(cf / (1 + rate_per_period) ** t for t, cf in enumerate(cashflows))


def _calculate_irr(cashflows: list[float], max_iterations: int = 1000) -> float | None:
    """Binary search IRR. Returns None if no solution found."""
    low, high = -0.999, 10.0
    for _ in range(max_iterations):
        mid = (low + high) / 2
        npv = sum(cf / (1 + mid) ** t for t, cf in enumerate(cashflows))
        if abs(npv) < 0.01:
            return mid
        if npv > 0:
            low = mid
        else:
            high = mid
    return None


class FinancialModelToolInput(BaseModel):
    periods: list[str] = Field(
        description="Ordered list of period name strings (e.g. ['Q1 2026', 'Q2 2026'])."
    )
    initiatives: list[dict] = Field(
        description=(
            "List of dicts with keys: id, title, period (name string), cost_gbp (float)."
        )
    )
    propositions: list[dict] = Field(
        description=(
            "List of dicts with keys: id, title, realisation_period (name string), "
            "annual_benefit_gbp (float)."
        )
    )
    discount_rate: float = Field(
        description="Annual discount rate as a decimal (e.g. 0.08 for 8%)."
    )
    period_duration_months: int = Field(
        description="Duration of each period in months (e.g. 3 for quarterly)."
    )
    filename: str = Field(
        description="Output filename (e.g. 'cost_benefit_model.xlsx'). "
        ".xlsx extension added automatically if missing."
    )
    agent_name: str = Field(
        description="Name of the agent producing this output (used for output tracking)."
    )


class FinancialModelTool(BaseTool):
    name: str = "FinancialModelTool"
    description: str = (
        "Build a 3-sheet financial model (.xlsx) with Cashflow Model, Financial Summary, "
        "and Assumptions sheets. Calculates NPV, IRR, payback period, and maximum borrowing "
        "requirement from initiative costs and proposition benefits. "
        "Returns the absolute file path."
    )
    args_schema: type[BaseModel] = FinancialModelToolInput
    slug: str

    def _run(
        self,
        periods: list[str],
        initiatives: list[dict],
        propositions: list[dict],
        discount_rate: float,
        period_duration_months: int,
        filename: str,
        agent_name: str,
    ) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return "Error: openpyxl not installed — run: pip install openpyxl"

        settings = get_settings()
        outputs_dir = Path(settings.projects_dir) / self.slug / "outputs"
        outputs_dir.mkdir(parents=True, exist_ok=True)

        if not filename.endswith(".xlsx"):
            filename = f"{filename}.xlsx"
        file_path = outputs_dir / filename

        try:
            # Pre-compute period index map
            period_index = {name: i for i, name in enumerate(periods)}
            n = len(periods)

            # Per-period benefit from annual_benefit_gbp, pro-rated by period_duration
            benefit_multiplier = period_duration_months / 12.0

            # Build cashflow arrays
            costs = [0.0] * n
            for init in initiatives:
                p = period_index.get(init["period"])
                if p is not None:
                    costs[p] += float(init.get("cost_gbp", 0))

            benefits = [0.0] * n
            for prop in propositions:
                r_idx = period_index.get(prop.get("realisation_period"))
                if r_idx is None:
                    continue
                per_period = float(prop.get("annual_benefit_gbp", 0)) * benefit_multiplier
                for p in range(r_idx, n):
                    benefits[p] += per_period

            net = [benefits[p] - costs[p] for p in range(n)]
            cumulative = []
            running = 0.0
            for v in net:
                running += v
                cumulative.append(running)

            # Payback: first period where cumulative turns positive
            payback_period = None
            for i, cum in enumerate(cumulative):
                if cum >= 0:
                    payback_period = periods[i]
                    break

            # Financial metrics
            rate_per_period = discount_rate * (period_duration_months / 12.0)
            npv = _calculate_npv(net, rate_per_period)
            irr = _calculate_irr(net)
            max_borrowing = min(cumulative)
            total_investment = sum(costs)
            total_benefits = sum(benefits)

            # ── Build workbook ────────────────────────────────────────────
            wb = openpyxl.Workbook()

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="1F397D")  # navy

            # ── Sheet 1: Cashflow Model ───────────────────────────────────
            ws1 = wb.active
            ws1.title = "Cashflow Model"

            # Header row
            cell = ws1.cell(row=1, column=1, value="Item")
            cell.font = header_font
            cell.fill = header_fill
            for col_i, period_name in enumerate(periods, start=2):
                cell = ws1.cell(row=1, column=col_i, value=period_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Initiative cost rows
            row = 2
            for init in initiatives:
                ws1.cell(row=row, column=1, value=f"Cost: {init['title']}")
                p = period_index.get(init["period"])
                if p is not None:
                    ws1.cell(row=row, column=p + 2, value=-float(init.get("cost_gbp", 0)))
                row += 1

            # Proposition benefit rows
            for prop in propositions:
                ws1.cell(row=row, column=1, value=f"Benefit: {prop['title']}")
                r_idx = period_index.get(prop.get("realisation_period"))
                if r_idx is not None:
                    per_period = float(prop.get("annual_benefit_gbp", 0)) * benefit_multiplier
                    for p in range(r_idx, n):
                        ws1.cell(row=row, column=p + 2, value=per_period)
                row += 1

            # Net Cashflow row
            net_label = ws1.cell(row=row, column=1, value="Net Cashflow")
            net_label.font = Font(bold=True)
            for col_i, v in enumerate(net, start=2):
                ws1.cell(row=row, column=col_i, value=v)
            row += 1

            # Cumulative Cashflow row
            cum_label = ws1.cell(row=row, column=1, value="Cumulative Cashflow")
            cum_label.font = Font(bold=True)
            for col_i, v in enumerate(cumulative, start=2):
                ws1.cell(row=row, column=col_i, value=v)

            # Auto-width columns
            for col in ws1.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            # ── Sheet 2: Financial Summary ────────────────────────────────
            ws2 = wb.create_sheet("Financial Summary")
            ws2.cell(row=1, column=1, value="Metric").font = header_font
            ws2.cell(row=1, column=1).fill = header_fill
            ws2.cell(row=1, column=2, value="Value").font = header_font
            ws2.cell(row=1, column=2).fill = header_fill
            summary_rows = [
                ("NPV (£)", round(npv, 2)),
                ("IRR", round(irr, 4) if irr is not None else None),
                ("Payback Period", payback_period),
                ("Maximum Borrowing Requirement (£)", round(max_borrowing, 2)),
                ("Total Investment (£)", round(total_investment, 2)),
                ("Total Benefits over Horizon (£)", round(total_benefits, 2)),
            ]
            for i, (label, value) in enumerate(summary_rows, start=2):
                ws2.cell(row=i, column=1, value=label)
                ws2.cell(row=i, column=2, value=value)
            ws2.column_dimensions["A"].width = 40
            ws2.column_dimensions["B"].width = 20

            # ── Sheet 3: Assumptions ──────────────────────────────────────
            ws3 = wb.create_sheet("Assumptions")
            ws3.cell(row=1, column=1, value="Parameter").font = header_font
            ws3.cell(row=1, column=1).fill = header_fill
            ws3.cell(row=1, column=2, value="Value").font = header_font
            ws3.cell(row=1, column=2).fill = header_fill
            assumption_rows = [
                ("Discount Rate", f"{discount_rate * 100:.1f}%"),
                ("Period Duration (months)", period_duration_months),
                ("Number of Periods", n),
                ("Benefit pro-ration multiplier", benefit_multiplier),
            ]
            for i, (k, v) in enumerate(assumption_rows, start=2):
                ws3.cell(row=i, column=1, value=k)
                ws3.cell(row=i, column=2, value=v)
            ws3.column_dimensions["A"].width = 35
            ws3.column_dimensions["B"].width = 20

            # Initiatives sub-table
            ws3.cell(row=7, column=1, value="Initiative Costs Used").font = Font(bold=True)
            ws3.cell(row=8, column=1, value="ID")
            ws3.cell(row=8, column=2, value="Title")
            ws3.cell(row=8, column=3, value="Period")
            ws3.cell(row=8, column=4, value="Cost (£)")
            for j, init in enumerate(initiatives, start=9):
                ws3.cell(row=j, column=1, value=init.get("id", ""))
                ws3.cell(row=j, column=2, value=init.get("title", ""))
                ws3.cell(row=j, column=3, value=init.get("period", ""))
                ws3.cell(row=j, column=4, value=init.get("cost_gbp", 0))

            prop_start = 9 + len(initiatives) + 1
            ws3.cell(row=prop_start, column=1, value="Proposition Benefits Used").font = Font(bold=True)
            ws3.cell(row=prop_start + 1, column=1, value="ID")
            ws3.cell(row=prop_start + 1, column=2, value="Title")
            ws3.cell(row=prop_start + 1, column=3, value="Realisation Period")
            ws3.cell(row=prop_start + 1, column=4, value="Annual Benefit (£)")
            for j, prop in enumerate(propositions, start=prop_start + 2):
                ws3.cell(row=j, column=1, value=prop.get("id", ""))
                ws3.cell(row=j, column=2, value=prop.get("title", ""))
                ws3.cell(row=j, column=3, value=prop.get("realisation_period", ""))
                ws3.cell(row=j, column=4, value=prop.get("annual_benefit_gbp", 0))

            wb.save(file_path)
            insert_agent_output_sync(
                slug=self.slug,
                agent_name=agent_name,
                output_type="xlsx",
                file_path=str(file_path),
            )
        except (OSError, ValueError) as e:
            return f"Error: render failed — {e}"

        return str(file_path)
