# tests/test_excel_output.py
import json
import pytest
from pathlib import Path
from unittest.mock import patch


@pytest.fixture(autouse=True)
def isolated_projects_dir(tmp_path, monkeypatch):
    """Redirect PROJECTS_DIR to a temp directory for each test."""
    monkeypatch.setenv("PROJECTS_DIR", str(tmp_path))
    from api.config import get_settings
    get_settings.cache_clear()
    yield tmp_path
    get_settings.cache_clear()


@pytest.fixture
def slug(isolated_projects_dir):
    slug = "excel-test-project"
    project_dir = isolated_projects_dir / slug
    outputs_dir = project_dir / "outputs"
    outputs_dir.mkdir(parents=True)
    # ExcelOutputTool needs a project row in SQLite — patch insert_agent_output_sync
    return slug


def test_excel_output_tool_writes_file(slug, isolated_projects_dir):
    """ExcelOutputTool writes an xlsx file with correct headers and rows."""
    from agents.tools.excel_output import ExcelOutputTool
    import openpyxl

    with patch("agents.tools.excel_output.insert_agent_output_sync"):
        tool = ExcelOutputTool(slug=slug)
        rows = [
            {"rank": 1, "title": "Alpha", "score": 90.0},
            {"rank": 2, "title": "Beta", "score": 75.0},
        ]
        result = tool._run(
            rows=rows,
            columns=["rank", "title", "score"],
            filename="portfolio_register.xlsx",
            agent_name="portfolio_manager",
        )

    file_path = Path(result)
    assert file_path.exists(), "XLSX file was not created"
    assert file_path.suffix == ".xlsx"

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Headers in row 1
    headers = [ws.cell(row=1, column=i).value for i in range(1, 4)]
    assert headers == ["rank", "title", "score"]

    # Data rows
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "Alpha"
    assert ws.cell(row=3, column=2).value == "Beta"


def test_excel_output_tool_returns_absolute_path(slug):
    """Return value is the absolute path to the written file."""
    from agents.tools.excel_output import ExcelOutputTool

    with patch("agents.tools.excel_output.insert_agent_output_sync"):
        tool = ExcelOutputTool(slug=slug)
        result = tool._run(
            rows=[{"col": "val"}],
            columns=["col"],
            filename="test.xlsx",
            agent_name="test_agent",
        )

    assert Path(result).is_absolute()
    assert result.endswith(".xlsx")


def test_excel_output_tool_appends_xlsx_extension(slug):
    """filename without .xlsx extension gets it added automatically."""
    from agents.tools.excel_output import ExcelOutputTool

    with patch("agents.tools.excel_output.insert_agent_output_sync"):
        tool = ExcelOutputTool(slug=slug)
        result = tool._run(
            rows=[{"x": 1}],
            columns=["x"],
            filename="no_extension",
            agent_name="test_agent",
        )

    assert result.endswith(".xlsx")
    assert Path(result).exists()


def test_excel_output_tool_header_is_bold(slug):
    """Header row cells have bold font."""
    from agents.tools.excel_output import ExcelOutputTool
    import openpyxl

    with patch("agents.tools.excel_output.insert_agent_output_sync"):
        tool = ExcelOutputTool(slug=slug)
        result = tool._run(
            rows=[{"name": "X"}],
            columns=["name"],
            filename="bold_test.xlsx",
            agent_name="test_agent",
        )

    wb = openpyxl.load_workbook(result)
    ws = wb.active
    assert ws.cell(row=1, column=1).font.bold is True


def test_excel_output_tool_freeze_panes(slug):
    """Freeze panes set to A2 (header row frozen)."""
    from agents.tools.excel_output import ExcelOutputTool
    import openpyxl

    with patch("agents.tools.excel_output.insert_agent_output_sync"):
        tool = ExcelOutputTool(slug=slug)
        result = tool._run(
            rows=[{"a": 1}],
            columns=["a"],
            filename="freeze_test.xlsx",
            agent_name="test_agent",
        )

    wb = openpyxl.load_workbook(result)
    ws = wb.active
    assert ws.freeze_panes == "A2"
