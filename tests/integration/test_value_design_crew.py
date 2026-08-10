# tests/integration/test_value_design_crew.py
"""
End-to-end integration test for the Value Design Crew.

Requires:
- ANTHROPIC_API_KEY in .env
- Discovery outputs seeded by seed_discovery_outputs fixture

Run with: pytest -m integration -v
Takes 3–8 minutes.
"""
import contextlib
import json
import sqlite3
import pytest
from pathlib import Path
from api.config import get_settings


@pytest.mark.integration
def test_value_design_crew_end_to_end(test_slug, project_id, seed_discovery_outputs):
    """
    Run the full Value Design Crew and verify all outputs are produced.
    Uses claude-haiku for both agents (test LLM override).
    HITL pauses are auto-responded via HITL_AUTO_RESPOND='approved' set in conftest.
    """
    from agents.llm import get_test_llm
    from agents.crews.value_design_crew import create_value_design_crew

    settings = get_settings()
    db_path = Path(settings.database_dir) / f"{test_slug}.db"

    # Create a crew_run record
    conn = sqlite3.connect(db_path)
    cur = conn.execute(
        "INSERT INTO crew_runs (project_id, crew_name, status, started_at)"
        " VALUES (?,?,?, CURRENT_TIMESTAMP)",
        (project_id, "value_design", "running"),
    )
    conn.commit()
    run_id = cur.lastrowid
    conn.close()

    # Build crew with cheap test LLM (both agents use Haiku in tests)
    llm = get_test_llm()
    crew = create_value_design_crew(
        slug=test_slug,
        run_id=run_id,
        sector="logistics",
        llm=llm,
    )

    result = crew.kickoff()
    assert result is not None

    # Mark run completed (in production, run_service does this)
    with contextlib.closing(sqlite3.connect(db_path)) as conn:
        conn.execute(
            "UPDATE crew_runs SET status='completed', finished_at=CURRENT_TIMESTAMP WHERE id=?",
            (run_id,),
        )
        conn.commit()

    outputs_dir = Path(settings.projects_dir) / test_slug / "outputs"

    # Anything SQLiteStateTool writes lands under a _vN suffix - the tool reports "Written
    # to propositions.json" while the file on disk is propositions_v1.json - so a declared
    # JSON output is resolved through the ledger. The .xlsx below keeps its bare name: it
    # comes from a different tool that does not version.
    from agents.tools._db import current_output_path

    # 1. propositions exists and is a valid JSON array
    propositions_path = current_output_path(test_slug, "propositions")
    assert propositions_path is not None, "propositions not created"
    propositions = json.loads(propositions_path.read_text())
    assert isinstance(propositions, list), "propositions.json is not a JSON array"
    assert len(propositions) >= 1, "propositions.json contains no propositions"
    first = propositions[0]
    assert "id" in first, "Proposition missing 'id' field"
    assert "title" in first, "Proposition missing 'title' field"
    assert "change_articulation" in first, "Proposition missing 'change_articulation' field"
    assert "value_estimate" in first, "Proposition missing 'value_estimate' field"
    assert first["value_estimate"] in ("High", "Medium", "Low"), \
        f"Invalid value_estimate: {first['value_estimate']}"

    # 2. portfolio_register exists and is a valid JSON array
    portfolio_path = current_output_path(test_slug, "portfolio_register")
    assert portfolio_path is not None, "portfolio_register not created"
    portfolio = json.loads(portfolio_path.read_text())
    assert isinstance(portfolio, list), "portfolio_register.json is not a JSON array"
    assert len(portfolio) >= 1, "portfolio_register.json contains no entries"
    first_item = portfolio[0]
    assert "rank" in first_item, "Portfolio entry missing 'rank' field"
    assert "total_score" in first_item, "Portfolio entry missing 'total_score' field"
    assert first_item["rank"] == 1, "First portfolio entry should have rank 1"

    # 3. portfolio_register.xlsx exists
    xlsx_path = outputs_dir / "portfolio_register.xlsx"
    assert xlsx_path.exists(), "portfolio_register.xlsx not created"

    # 4. XLSX has correct structure
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    assert "rank" in headers, "portfolio_register.xlsx missing 'rank' column"
    assert "title" in headers, "portfolio_register.xlsx missing 'title' column"

    # 5. HITL records created
    with contextlib.closing(sqlite3.connect(db_path)) as conn:
        cur = conn.execute(
            "SELECT COUNT(*) FROM human_reviews WHERE crew_run_id=?", (run_id,)
        )
        hitl_count = cur.fetchone()[0]
    assert hitl_count >= 1, "No HITL reviews created during Value Design crew run"

    # 6. agent_outputs records created
    with contextlib.closing(sqlite3.connect(db_path)) as conn:
        cur = conn.execute(
            "SELECT DISTINCT agent_name FROM agent_outputs WHERE project_id=?",
            (project_id,),
        )
        agent_names = {row[0] for row in cur.fetchall()}
    assert "portfolio_manager" in agent_names, "Portfolio Manager produced no tracked output"
