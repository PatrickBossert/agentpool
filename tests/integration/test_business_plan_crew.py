# tests/integration/test_business_plan_crew.py
"""
End-to-end integration test for the Business Plan Crew.

Requires:
- ANTHROPIC_API_KEY in .env
- All prior crew outputs seeded by seed_delivery_outputs (which chains back through
  seed_architecture_outputs → seed_value_design_outputs → seed_discovery_outputs)

Run with: pytest -m integration -v
Takes 5-12 minutes.
"""
import contextlib
import json
import sqlite3
import pytest
from pathlib import Path
from api.config import get_settings


@pytest.mark.integration
def test_business_plan_crew_end_to_end(test_slug, project_id, seed_delivery_outputs):
    """
    Run the full Business Plan Crew and verify all three artefacts are produced.
    Uses claude-haiku for the agent (test LLM override).
    HITL pauses are auto-responded via HITL_AUTO_RESPOND='approved' set in conftest.
    """
    from agents.llm import get_test_llm
    from agents.crews.business_plan_crew import create_business_plan_crew

    settings = get_settings()
    db_path = Path(settings.database_dir) / f"{test_slug}.db"

    # Create a crew_run record
    conn = sqlite3.connect(db_path)
    cur = conn.execute(
        "INSERT INTO crew_runs (project_id, crew_name, status, started_at)"
        " VALUES (?,?,?, CURRENT_TIMESTAMP)",
        (project_id, "business_plan", "running"),
    )
    conn.commit()
    run_id = cur.lastrowid
    conn.close()

    llm = get_test_llm()
    crew = create_business_plan_crew(
        slug=test_slug,
        run_id=run_id,
        llm_mode="standard",
        sector="logistics",
        llm=llm,
    )

    result = crew.kickoff()
    assert result is not None

    # Mark run completed
    with contextlib.closing(sqlite3.connect(db_path)) as conn:
        conn.execute(
            "UPDATE crew_runs SET status='completed', finished_at=CURRENT_TIMESTAMP WHERE id=?",
            (run_id,),
        )
        conn.commit()

    outputs_dir = Path(settings.projects_dir) / test_slug / "outputs"

    # 1. business_plan.docx exists and is non-empty
    docx_path = outputs_dir / "business_plan.docx"
    assert docx_path.exists(), "business_plan.docx not created"
    assert docx_path.stat().st_size > 0, "business_plan.docx is empty"

    # 2. executive_presentation.pptx exists and is non-empty
    pptx_path = outputs_dir / "executive_presentation.pptx"
    assert pptx_path.exists(), "executive_presentation.pptx not created"
    assert pptx_path.stat().st_size > 0, "executive_presentation.pptx is empty"

    # 3. cost_benefit_model.xlsx exists with 3 sheets and numeric NPV + max_borrowing
    import openpyxl
    xlsx_path = outputs_dir / "cost_benefit_model.xlsx"
    assert xlsx_path.exists(), "cost_benefit_model.xlsx not created"
    wb = openpyxl.load_workbook(xlsx_path)
    assert len(wb.sheetnames) == 3, f"Expected 3 sheets, got {wb.sheetnames}"

    ws_summary = wb["Financial Summary"]
    npv_value = None
    max_borrow = None
    for row in ws_summary.iter_rows(values_only=True):
        if row[0] and "NPV" in str(row[0]):
            npv_value = row[1]
        if row[0] and "Maximum Borrowing" in str(row[0]):
            max_borrow = row[1]
    assert isinstance(npv_value, (int, float)), f"NPV not numeric: {npv_value}"
    assert isinstance(max_borrow, (int, float)), f"Max borrowing not numeric: {max_borrow}"

    # 4. At least 2 HITL records (context gate + review gate)
    with contextlib.closing(sqlite3.connect(db_path)) as conn:
        cur = conn.execute(
            "SELECT COUNT(*) FROM human_reviews WHERE crew_run_id=?", (run_id,)
        )
        hitl_count = cur.fetchone()[0]
    assert hitl_count >= 2, f"Expected at least 2 HITL reviews, got {hitl_count}"

    # 5. agent_outputs record for business_plan_generator
    with contextlib.closing(sqlite3.connect(db_path)) as conn:
        cur = conn.execute(
            "SELECT DISTINCT agent_name FROM agent_outputs WHERE project_id=?",
            (project_id,),
        )
        agent_names = {row[0] for row in cur.fetchall()}
    assert "business_plan_generator" in agent_names, \
        "Business Plan Generator produced no tracked output"
