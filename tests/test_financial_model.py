# tests/test_financial_model.py
import pytest
from pathlib import Path
from unittest.mock import patch


@pytest.fixture(autouse=True)
def isolated_projects_dir(tmp_path, monkeypatch):
    monkeypatch.setenv("PROJECTS_DIR", str(tmp_path))
    from api.config import get_settings
    get_settings.cache_clear()
    yield tmp_path
    get_settings.cache_clear()


@pytest.fixture
def slug(isolated_projects_dir):
    slug = "financial-model-test"
    outputs_dir = isolated_projects_dir / slug / "outputs"
    outputs_dir.mkdir(parents=True)
    return slug


@pytest.fixture
def sample_inputs():
    return {
        "periods": ["Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"],
        "initiatives": [
            {"id": "INIT-001", "title": "Automate Order Entry", "period": "Q1 2026", "cost_gbp": 100000.0},
            {"id": "INIT-002", "title": "Integrate WMS", "period": "Q2 2026", "cost_gbp": 200000.0},
        ],
        "propositions": [
            {
                "id": "VP-001",
                "title": "Automated Order Management",
                "realisation_period": "Q2 2026",
                "annual_benefit_gbp": 500000.0,
            },
            {
                "id": "VP-002",
                "title": "Integrated Supply Chain",
                "realisation_period": "Q3 2026",
                "annual_benefit_gbp": 200000.0,
            },
        ],
        "discount_rate": 0.08,
        "period_duration_months": 3,
        "filename": "cost_benefit_model.xlsx",
        "agent_name": "business_plan_generator",
    }


def test_financial_model_writes_file(slug, sample_inputs):
    """FinancialModelTool creates an .xlsx file at the returned path."""
    from agents.tools.financial_model import FinancialModelTool
    with patch("agents.tools.financial_model.insert_agent_output_sync"):
        tool = FinancialModelTool(slug=slug)
        result = tool._run(**sample_inputs)
    assert Path(result).exists()


def test_financial_model_returns_absolute_path(slug, sample_inputs):
    """Return value is an absolute path ending in .xlsx."""
    from agents.tools.financial_model import FinancialModelTool
    with patch("agents.tools.financial_model.insert_agent_output_sync"):
        tool = FinancialModelTool(slug=slug)
        result = tool._run(**sample_inputs)
    assert Path(result).is_absolute()
    assert result.endswith(".xlsx")


def test_financial_model_has_three_sheets(slug, sample_inputs):
    """Workbook contains exactly 3 sheets: Cashflow Model, Financial Summary, Assumptions."""
    import openpyxl
    from agents.tools.financial_model import FinancialModelTool
    with patch("agents.tools.financial_model.insert_agent_output_sync"):
        tool = FinancialModelTool(slug=slug)
        result = tool._run(**sample_inputs)
    wb = openpyxl.load_workbook(result)
    assert len(wb.sheetnames) == 3
    assert "Cashflow Model" in wb.sheetnames
    assert "Financial Summary" in wb.sheetnames
    assert "Assumptions" in wb.sheetnames


def test_financial_model_npv_is_float(slug, sample_inputs):
    """Financial Summary sheet contains a numeric NPV value."""
    import openpyxl
    from agents.tools.financial_model import FinancialModelTool
    with patch("agents.tools.financial_model.insert_agent_output_sync"):
        tool = FinancialModelTool(slug=slug)
        result = tool._run(**sample_inputs)
    wb = openpyxl.load_workbook(result)
    ws = wb["Financial Summary"]
    npv_value = None
    for row in ws.iter_rows(values_only=True):
        if row[0] and "NPV" in str(row[0]):
            npv_value = row[1]
            break
    assert npv_value is not None
    assert isinstance(npv_value, (int, float))


def test_financial_model_irr_is_float_or_none(slug, sample_inputs):
    """IRR in Financial Summary is a float (or None if no solution)."""
    import openpyxl
    from agents.tools.financial_model import FinancialModelTool
    with patch("agents.tools.financial_model.insert_agent_output_sync"):
        tool = FinancialModelTool(slug=slug)
        result = tool._run(**sample_inputs)
    wb = openpyxl.load_workbook(result)
    ws = wb["Financial Summary"]
    irr_value = None
    found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] and "IRR" in str(row[0]):
            irr_value = row[1]
            found = True
            break
    assert found, "IRR row not found in Financial Summary"
    assert irr_value is None or isinstance(irr_value, (int, float))


def test_financial_model_max_borrowing_is_nonpositive(slug, sample_inputs):
    """Max Borrowing Requirement is <= 0 (peak outflow before returns begin)."""
    import openpyxl
    from agents.tools.financial_model import FinancialModelTool
    with patch("agents.tools.financial_model.insert_agent_output_sync"):
        tool = FinancialModelTool(slug=slug)
        result = tool._run(**sample_inputs)
    wb = openpyxl.load_workbook(result)
    ws = wb["Financial Summary"]
    max_borrow = None
    for row in ws.iter_rows(values_only=True):
        if row[0] and "Maximum Borrowing" in str(row[0]):
            max_borrow = row[1]
            break
    assert max_borrow is not None
    assert max_borrow <= 0


def test_financial_model_appends_xlsx_extension(slug, sample_inputs):
    """filename without .xlsx extension gets it added automatically."""
    from agents.tools.financial_model import FinancialModelTool
    inputs = {**sample_inputs, "filename": "cost_benefit_no_ext"}
    with patch("agents.tools.financial_model.insert_agent_output_sync"):
        tool = FinancialModelTool(slug=slug)
        result = tool._run(**inputs)
    assert result.endswith(".xlsx")
    assert Path(result).exists()


def test_financial_model_error_on_write_failure(slug, sample_inputs):
    """Returns error string when the file cannot be saved."""
    import openpyxl
    from agents.tools.financial_model import FinancialModelTool
    with patch("agents.tools.financial_model.insert_agent_output_sync"), \
         patch.object(openpyxl.Workbook, "save", side_effect=OSError("disk full")):
        tool = FinancialModelTool(slug=slug)
        result = tool._run(**sample_inputs)
    assert result.startswith("Error: render failed")
