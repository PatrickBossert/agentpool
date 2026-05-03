import shutil
import pytest
from pathlib import Path
from api.config import get_settings
from api.database import get_connection, insert_agent_output, fetch_project

SLUG = "bp-test"
PROJECT = {
    "client_slug": SLUG,
    "llm_mode": "standard",
    "sector": "rail",
}


@pytest.fixture(autouse=True)
def clean():
    settings = get_settings()
    db_path = Path(settings.database_dir) / f"{SLUG}.db"
    proj_dir = Path(settings.projects_dir) / SLUG
    db_path.unlink(missing_ok=True)
    if proj_dir.exists():
        shutil.rmtree(proj_dir)
    yield
    get_settings.cache_clear()
    db_path.unlink(missing_ok=True)
    if proj_dir.exists():
        shutil.rmtree(proj_dir)


def _write_excel_file(path: Path) -> None:
    """Write a minimal XLSX with the Financial Summary sheet that FinancialModelTool produces."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Cashflow Model"
    ws2 = wb.create_sheet("Financial Summary")
    rows = [
        ("NPV (£)", 4_200_000.0),
        ("IRR", 0.234),
        ("Payback Period", "Q3 2026"),
        ("Maximum Borrowing Requirement (£)", -1_100_000.0),
        ("Total Investment (£)", 2_800_000.0),
        ("Total Benefits over Horizon (£)", 9_600_000.0),
    ]
    for i, (label, value) in enumerate(rows, start=2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=value)
    wb.save(path)


async def _insert_excel_output(file_path: str) -> int:
    async with get_connection(SLUG) as conn:
        project = await fetch_project(conn, slug=SLUG)
        return await insert_agent_output(
            conn,
            project_id=project["id"],
            agent_name="test_agent",
            output_type="excel",
            file_path=file_path,
            version=1,
        )


@pytest.mark.asyncio
async def test_get_financial_summary_returns_metrics(client):
    """Create project + write XLSX + insert row → GET returns 200 with all 6 keys."""
    await client.post("/projects", json=PROJECT)
    settings = get_settings()
    outputs_dir = Path(settings.projects_dir) / SLUG / "outputs"
    outputs_dir.mkdir(parents=True, exist_ok=True)
    xlsx_file = outputs_dir / "cost_benefit_model.xlsx"
    _write_excel_file(xlsx_file)
    await _insert_excel_output(str(xlsx_file))

    resp = await client.get(f"/projects/{SLUG}/financial-summary")
    assert resp.status_code == 200
    data = resp.json()
    assert data["npv"] == pytest.approx(4_200_000.0)
    assert data["irr"] == pytest.approx(0.234)
    assert data["payback_period"] == "Q3 2026"
    assert data["max_borrowing"] == pytest.approx(-1_100_000.0)
    assert data["total_investment"] == pytest.approx(2_800_000.0)
    assert data["total_benefits"] == pytest.approx(9_600_000.0)


@pytest.mark.asyncio
async def test_get_financial_summary_unknown_project_404(client):
    resp = await client.get("/projects/ghost-project/financial-summary")
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_get_financial_summary_no_output_404(client):
    """Valid project with no excel output row → 404."""
    await client.post("/projects", json=PROJECT)
    resp = await client.get(f"/projects/{SLUG}/financial-summary")
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_get_financial_summary_missing_sheet_404(client):
    """XLSX exists but has no Financial Summary sheet → 404."""
    import openpyxl
    await client.post("/projects", json=PROJECT)
    settings = get_settings()
    outputs_dir = Path(settings.projects_dir) / SLUG / "outputs"
    outputs_dir.mkdir(parents=True, exist_ok=True)
    xlsx_file = outputs_dir / "bad_model.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"  # no Financial Summary sheet
    wb.save(xlsx_file)
    await _insert_excel_output(str(xlsx_file))

    resp = await client.get(f"/projects/{SLUG}/financial-summary")
    assert resp.status_code == 404
